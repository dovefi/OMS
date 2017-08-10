# -*- coding=utf-8 -*-
import sys, os
import openpyxl

source = "test.xlsx"
sheet_name = "Sheet1"

USAGE = """\
    请使用以下的格式填写excel,
    第一行 必须为:学校  域名
         A             B
    1 | 学校      |    域名
    2 | 广州大学   |    gzdx
    3 | 北京大学   |    bjdx
    
    """
CONF_TEMPLATE = """\
server {
        listen      80;
        server_name  DOMAIN.dyfchk2.kuxiao.cn;
        charset      utf-8;

        location / {
            proxy_pass   http://127.0.0.1:4673;
            proxy_set_header Host $host;
            proxy_set_header X-Forward-For $remote_addr;
            proxy_set_header X-Real-IP $remote_addr;
        }
        location = /crossdomain.xml {
            root /etc/nginx/conf.d/dyf.d/domain;
        }

        location = /proxy.html {
            root /etc/nginx/conf.d/dyf.d/domain;
        }

        location = /xdomain.min.js {
            root /etc/nginx/conf.d/dyf.d/domain;
        }

        location = /art/api/loadMarkdown {
            proxy_pass   http://127.0.0.1:1915;
            proxy_set_header Host $host;
            proxy_set_header X-Forward-For $remote_addr;
            proxy_set_header X-Real-IP $remote_addr;
        }
}
"""

if __name__ == '__main__':

    workbook = openpyxl.load_workbook(source)
    sheet = workbook.get_sheet_by_name(sheet_name)

    if sheet.max_column != 2:
        print(USAGE)
        sys.exit()
    start_row = 1
    start_column = 2

    if sheet['A1'].value == "学校" and sheet['B1'].value == "域名":
        start_row = 2
    else:
        print(USAGE)
        sys.exit()

    if os.path.isdir("conf") is False:
        os.mkdir('./conf')

    for c in range(start_column, sheet.max_column+1):
        for r in range(start_row, sheet.max_row+1):
            if sheet.cell(row=r, column=c).value is not None:
                f = open("./conf/" + str(sheet.cell(row=r, column=c).value) + ".conf", 'w', encoding='utf-8')
                f.write(CONF_TEMPLATE.replace("DOMAIN", str(sheet.cell(row=r, column=c).value)))
                f.close()
            else:
                log = open("./out.log", 'w+', encoding='utf-8')
                log.write("第{0}列，第{1}行的域名为空" .format(c,r))
                log.close()
                print("第{0}列，第{1}行的域名为空" .format(c,r))







